import os

import openpyxl


class TestReportExporter:

    @staticmethod
    def export(report_path):
        """
        テンプレート(Excelファイル)からExcel形式の帳票を出力する
        :param report_path: 帳票ファイルの出力先パス
        :return:
        """

        # テンプレートファイルを読み込み
        template_workbook = openpyxl.load_workbook(os.path.join("template", "test-report.xlsx"))

        try:
            # シートを取得
            worksheet = template_workbook.worksheets[0]

            # 可変項目を書き込み
            worksheet["B3"] = "データA"
            base_row_index = 6
            detail_list = [
                {"a": "1", "b": "データA", "c": "これはデータAです"},
                {"a": "2", "b": "データB", "c": "これはデータBです"},
                {"a": "3", "b": "データC", "c": "これはデータCです"}
            ]
            worksheet.cell(row=base_row_index, column=1, value=detail_list[0].get("a"))
            worksheet.cell(row=base_row_index, column=2, value=detail_list[0].get("b"))
            worksheet.cell(row=base_row_index, column=3, value=detail_list[0].get("c"))
            worksheet.cell(row=base_row_index + 1, column=1, value=detail_list[1].get("a"))
            worksheet.cell(row=base_row_index + 1, column=2, value=detail_list[1].get("b"))
            worksheet.cell(row=base_row_index + 1, column=3, value=detail_list[1].get("c"))
            worksheet.cell(row=base_row_index + 2, column=1, value=detail_list[2].get("a"))
            worksheet.cell(row=base_row_index + 2, column=2, value=detail_list[2].get("b"))
            worksheet.cell(row=base_row_index + 2, column=3, value=detail_list[2].get("c"))

            # 帳票を保存
            template_workbook.save(report_path)
        finally:
            if template_workbook is not None:
                template_workbook.close()
